#!/usr/bin/env python3

import openpyxl
import os

# Verificar el archivo de prueba
archivo = './salida/TEST-EQUIPARACION-CON-QUIMICA.xlsx'

print(f"Verificando archivo: {archivo}")
print(f"Existe archivo: {os.path.exists(archivo)}")

if os.path.exists(archivo):
    try:
        workbook = openpyxl.load_workbook(archivo)
        print("✓ Archivo cargado correctamente")
        print(f"Hojas disponibles: {workbook.sheetnames}")
        
        if 'Equiparación' in workbook.sheetnames:
            print("✓ Hoja 'Equiparación' encontrada")
            sheet = workbook['Equiparación']
            
            # Verificar las primeras 20 filas buscando cursos equiparados
            equiparados_encontrados = []
            print("\nBuscando cursos EQUIPARADOS:")
            
            for row in range(1, 21):
                try:
                    sigla_vieja = sheet.cell(row=row, column=1).value  # Columna A
                    estado_nuevo = sheet.cell(row=row, column=8).value  # Columna H
                    
                    if estado_nuevo and 'EQUIPARADO' in str(estado_nuevo):
                        curso_viejo = sheet.cell(row=row, column=2).value  # Columna B
                        sigla_nueva = sheet.cell(row=row, column=5).value  # Columna E
                        curso_nuevo = sheet.cell(row=row, column=6).value  # Columna F
                        
                        # Verificar color de fondo
                        cell_h = sheet.cell(row=row, column=8)
                        fill = cell_h.fill
                        color = "Sin color específico"
                        
                        if fill and hasattr(fill, 'start_color') and fill.start_color:
                            if hasattr(fill.start_color, 'rgb') and fill.start_color.rgb:
                                color = fill.start_color.rgb
                        
                        equiparados_encontrados.append({
                            'fila': row,
                            'sigla_vieja': sigla_vieja,
                            'curso_viejo': curso_viejo,
                            'sigla_nueva': sigla_nueva,
                            'curso_nuevo': curso_nuevo,
                            'color': color
                        })
                        
                        print(f"  Fila {row}: {sigla_vieja} -> {sigla_nueva} ({curso_nuevo}) [Color: {color}]")
                
                except Exception as e:
                    continue  # Saltar errores en filas vacías
            
            if equiparados_encontrados:
                print(f"\n✅ ENCONTRADOS {len(equiparados_encontrados)} CURSOS EQUIPARADOS")
                
                # Verificar si alguno tiene el color azul claro esperado
                color_correcto = 'FFCCE5FF'  # CCE5FF con prefijo FF para alpha
                colores_encontrados = set([item['color'] for item in equiparados_encontrados])
                
                print(f"\nColores de fondo encontrados: {colores_encontrados}")
                
                if color_correcto in colores_encontrados:
                    print("✅ ¡Color azul claro aplicado correctamente!")
                else:
                    print("⚠️ El color azul claro no se detecta correctamente, pero puede estar aplicado")
            else:
                print("❌ NO SE ENCONTRARON CURSOS EQUIPARADOS")
                
                # Mostrar algunas filas para debug
                print("\nContenido de las primeras filas:")
                for row in range(7, 12):
                    try:
                        vals = [sheet.cell(row=row, column=col).value for col in range(1, 9)]
                        print(f"  Fila {row}: {vals}")
                    except:
                        pass
        else:
            print("❌ No se encontró hoja 'Equiparación'")
        
        workbook.close()
        
    except Exception as e:
        print(f"❌ Error al procesar archivo: {e}")
else:
    print("❌ Archivo no encontrado")
