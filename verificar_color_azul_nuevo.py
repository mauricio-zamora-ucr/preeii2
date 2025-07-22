#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para verificar que los cursos EQUIPARADO ahora usan el color azul claro (CCE5FF)
en vez del color verde anterior (D5E8D4)
"""

from openpyxl import load_workbook
import os

def verificar_color_equiparado(archivo_excel):
    """
    Verifica que los cursos EQUIPARADO tengan el color azul claro correcto
    """
    try:
        wb = load_workbook(archivo_excel)
        
        if 'Equiparación' not in wb.sheetnames:
            return f"❌ No se encontró la hoja 'Equiparación' en {archivo_excel}"
        
        sheet = wb['Equiparación']
        colores_encontrados = set()
        cursos_equiparados = []
        
        # Buscar en la columna H (nuevo plan) por cursos EQUIPARADO
        for row in range(3, sheet.max_row + 1):
            estado_cell = sheet[f'H{row}']
            if estado_cell.value and 'EQUIPARADO' in str(estado_cell.value).upper():
                # Verificar el color de fondo
                if estado_cell.fill.start_color.index and estado_cell.fill.start_color.index != '00000000':
                    color_hex = estado_cell.fill.start_color.index
                    colores_encontrados.add(color_hex)
                    
                    # Obtener información del curso
                    codigo_cell = sheet[f'A{row}']
                    nombre_cell = sheet[f'B{row}']
                    
                    curso_info = {
                        'codigo': codigo_cell.value if codigo_cell.value else '',
                        'nombre': nombre_cell.value if nombre_cell.value else '',
                        'estado': estado_cell.value,
                        'color': color_hex,
                        'fila': row
                    }
                    cursos_equiparados.append(curso_info)
        
        return {
            'archivo': os.path.basename(archivo_excel),
            'cursos_equiparados': cursos_equiparados,
            'colores_encontrados': colores_encontrados,
            'color_esperado': ['FFCCE5FF', '00CCE5FF'],  # Ambos formatos válidos para azul claro
            'cambio_exitoso': any(color in colores_encontrados for color in ['FFCCE5FF', '00CCE5FF']) and '00D5E8D4' not in colores_encontrados
        }
        
    except Exception as e:
        return f"❌ Error procesando {archivo_excel}: {str(e)}"

def main():
    """
    Función principal para verificar el cambio de color
    """
    # Archivos a verificar (selección de algunos estudiantes)
    archivos_test = [
        '/workspaces/preeii2/salida/B36447-ARIEL SANCHEZ MONTEALEGRE.xlsx',
        '/workspaces/preeii2/salida/A01776-CARLOS MAURICIO GOMEZ GARCIA.xlsx',  
        '/workspaces/preeii2/salida/C14407-ASAEL MAITA GARCIA.xlsx'
    ]
    
    print("🔍 Verificando el cambio de color de EQUIPARADO (verde → azul claro)")
    print("=" * 70)
    
    for archivo in archivos_test:
        if os.path.exists(archivo):
            resultado = verificar_color_equiparado(archivo)
            
            if isinstance(resultado, dict):
                print(f"\n📁 Archivo: {resultado['archivo']}")
                print(f"📊 Cursos equiparados encontrados: {len(resultado['cursos_equiparados'])}")
                
                if resultado['cursos_equiparados']:
                    for curso in resultado['cursos_equiparados']:
                        color_status = "✅" if curso['color'] in ['FFCCE5FF', '00CCE5FF'] else "❌"
                        print(f"   {color_status} {curso['codigo']} - {curso['estado']} (Color: {curso['color']})")
                
                print(f"🎨 Colores encontrados: {resultado['colores_encontrados']}")
                print(f"🔄 Cambio exitoso: {'✅ SÍ' if resultado['cambio_exitoso'] else '❌ NO'}")
                
                # Verificar específicamente el color esperado
                if any(color in resultado['colores_encontrados'] for color in ['FFCCE5FF', '00CCE5FF']):
                    print("🎉 ¡Color azul claro (CCE5FF) confirmado!")
                elif '00D5E8D4' in resultado['colores_encontrados']:
                    print("⚠️  Todavía se detecta el color verde anterior (D5E8D4)")
                else:
                    print("❓ Color inesperado encontrado")
            else:
                print(f"\n❌ {archivo}: {resultado}")
        else:
            print(f"\n❌ Archivo no encontrado: {archivo}")
    
    print("\n" + "=" * 70)
    print("Verificación completada")

if __name__ == "__main__":
    main()
