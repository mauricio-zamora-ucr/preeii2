"""
Analizador de equiparación para el nuevo plan de estudios.
Analiza expedientes existentes y agrega hoja de equiparación con análisis específico.
"""

from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any
import xlsxwriter
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from dataclasses import dataclass

from ...domain.entities.expediente import Expediente
from ...domain.entities.curso import Curso


@dataclass
class AnalisisEquiparacion:
    """Resultado del análisis de equiparación para un expediente."""
    quimica_general_completa: bool = False
    quimica_intensiva_completa: bool = False
    precalculo_aprobado: bool = False
    cursos_equivalentes: List[Tuple[str, str]] = None
    recomendaciones: List[str] = None
    
    def __post_init__(self):
        if self.cursos_equivalentes is None:
            self.cursos_equivalentes = []
        if self.recomendaciones is None:
            self.recomendaciones = []


class EquiparacionAnalyzer:
    """Analizador de equiparación para transición al nuevo plan de estudios."""
    
    def __init__(self):
        # Tabla completa de equivalencias basada en el archivo Equiparacion.xlsx
        self.tabla_equivalencias = [
            # [Sigla_Vieja, Curso_Viejo, Creditos_Viejo, Sigla_Nueva, Curso_Nuevo, Creditos_Nuevo]
            ['EF-D', 'ACTIVIDAD DEPORTIVA', '0', 'EF-D', 'ACTIVIDAD DEPORTIVA', '0'],
            ['EG-CA', 'CURSO DE ARTE', '2', 'EG-CA', 'CURSO DE ARTE', '2'],
            ['EG-I', 'CURSO INTEGRADO DE HUMANIDADES I', '6', 'EG-I', 'CURSO INTEGRADO DE HUMANIDADES I', '6'],
            ['MA0001', 'PRECÁLCULO', '0', 'MA0001', 'PRECÁLCULO', '0'],
            ['MA1001', 'CÁLCULO I', '3', 'MA1001', 'CÁLCULO I', '3'],
            ['QU0100', 'QUÍMICA GENERAL I', '3', 'QU0114', 'QUIMICA GENERAL INTENSIVA', '4'],
            ['QU0101', 'LABORATORIO DE QUÍMICA GENERAL I', '1', 'QU0115', 'LABORATORIO QUIMICA GENERAL INTENSIVA', '1'],
            ['RP-1', 'REPERTORIO', '3', 'RP-1', 'REPERTORIO', '3'],
            ['EG-II', 'CURSO INTEGRADO DE HUMANIDADES II', '6', 'EG-II', 'CURSO INTEGRADO DE HUMANIDADES II', '6'],
            ['FS0210', 'FÍSICA GENERAL I', '3', 'FS0210', 'FÍSICA GENERAL I', '3'],
            ['FS0211', 'LABORATORIO DE FÍSICA GENERAL I', '1', 'FS0211', 'LABORATORIO DE FÍSICA GENERAL I', '1'],
            ['II0201', 'INTRODUCCIÓN A LA INGENIERÍA INDUSTRIAL', '2', 'II-1010', 'INTRODUCCIÓN A LA INGENIERÍA INDUSTRIAL', '2'],
            ['MA1002', 'CÁLCULO II', '4', 'MA1002', 'CÁLCULO II', '4'],
            ['QU0102', 'QUÍMICA GENERAL II', '3', 'QU-0114', 'QUIMICA GENERAL INTENSIVA', ''],
            ['QU0103', 'LABORATORIO DE QUÍMICA GENERAL II', '1', 'QU-0115', 'LABORATORIO QUIMICA GENERAL INTENSIVA', ''],
            ['CI0202', 'PRINCIPIOS DE INFORMÁTICA', '4', 'II1011', 'FUNDAMENTOS PARA TECNOLOGÍAS DIGITALES', '4'],
            ['FS0310', 'FÍSICA GENERAL II', '3', 'FS0310', 'FÍSICA GENERAL II', '3'],
            ['FS0311', 'LABORATORIO DE FÍSICA GENERAL II', '1', 'FS0311', 'LABORATORIO DE FÍSICA GENERAL II', '1'],
            ['II0306', 'PROBABILIDAD Y ESTADÍSTICA', '3', 'II1012', 'ESTADISTICA I PARA II', '3'],
            ['MA1003', 'CÁLCULO III', '4', 'MA1003', 'CÁLCULO III', '4'],
            ['MA1004', 'ÁLGEBRA LINEAL', '3', 'MA-1004', 'ALGEBRA LINEAL', '3'],
            ['FS0410', 'FÍSICA GENERAL III', '3', '', 'No tiene equivalencia en el nuevo plan de estudios', ''],
            ['FS0411', 'LABORATORIO DE FÍSICA GENERAL III', '1', '', 'No tiene equivalencia en el nuevo plan de estudios', ''],
            ['II0401', 'INVESTIGACIÓN DE OPERACIONES', '3', 'II2020', 'MODELOS DE OPTIMIZACIÓN INDUSTRIAL', '3'],
            ['II0402', 'INGENIERÍA DE CALIDAD I', '2', 'II2021', 'ESTADISTICA II PARA II', '4'],
            ['II0501', 'TECNOLOGÍAS DE INFORMACIÓN', '2', 'II3031', 'INGENIERIA DE LA INFORMACION', '3'],
            ['IM0202', 'DIBUJO I', '3', 'IM0101', 'GRÁFICA', '3'],
            ['MA1005', 'ECUACIONES DIFERENCIALES', '4', 'MA1005', 'ECUACIONES DIFERENCIALES ', '4'],
            ['IE0303', 'ELECTROTECNIA I', '3', 'II4042', 'FUNDAMENTOS PARA MANUFACTURA', '4'],
            ['II0302', 'DISEÑO DEL TRABAJO E INGENIERÍA DE FACTORES HUMANOS', '3', 'II3034', 'DISEÑO Y MEDICION DEL TRABAJO', '3'],
            ['II0502', 'INGENIERÍA DE CALIDAD II', '4', 'II4050', 'INGENIERIA DE CALIDAD Y MEJORA CONTINUA', '4'],
            ['II0503', 'SIMULACIÓN', '3', 'II-3030', 'SIMULACION Y SISTEMAS DINAMICOS', '3'],
            ['II0504', 'ADMINISTRACIÓN FINANCIERA Y CONTABLE I', '2', 'II2022', 'INGENIERIA ECONOMICA INDUSTRIAL I', '3'],
            ['IM0207', 'MECÁNICA I', '3', 'IM0207', 'MECÁNICA', '3'],
            ['II0601', 'GESTIÓN DE CALIDAD', '4', 'II5053', 'GERENCIA Y SISTEMAS DE GESTION INTEGRADOS', '3'],
            ['II0603', 'SISTEMAS AUTOMATIZADOS DE MANUFACTURA', '3', 'II4049', 'SISTEMAS DE MANUFACTURA', '4'],
            ['II0604', 'ADMINISTRACIÓN FINANCIERA Y CONTABLE II', '2', '', 'No tiene equivalencia en el nuevo plan de estudios', ''],
            ['II0605', 'LOGÍSTICA DE LA CADENA DEL VALOR I', '3', 'II4041', 'INGENIERIA CADENA DE SUMINISTRO I', '3'],
            ['II0606', 'TERMOFLUIDOS', '3', 'II4042', 'FUNDAMENTOS PARA MANUFACTURA', ''],
            ['II0701', 'DISEÑO DE SISTEMAS DE INFORMACIÓN', '3', 'II3037', 'ANALITICA INDUSTRIAL', '3'],
            ['II0602', 'DISEÑO DE EXPERIMENTOS', '3', 'II2023', 'ESTADISTICA III PARA II', '3'],
            ['II0702', 'COMPORTAMIENTO ORGANIZACIONAL', '2', 'II1013', 'GESTION DE LA INGENIERIA', '2'],
            ['II0703', 'INGENIERÍA DE OPERACIONES', '4', 'II4048', 'INGENIERIA DE OPERACIONES', '4'],
            ['II0704', 'INGENIERÍA ECONÓMICA Y FINANCIERA', '3', 'II2025', 'INGENIERIA ECONOMICA INDUSTRIAL II', '3'],
            ['II0705', 'LOGÍSTICA DE LA CADENA DEL VALOR II', '4', 'II4047', 'INGENIERIA CADENA DE SUMINISTRO II', '3'],
            ['SR-I', 'SEMINARIO DE REALIDAD NACIONAL I', '2', 'SR-I', 'SEMINARIO DE REALIDAD NACIONAL I', '2'],
            ['II0802', 'INGENIERÍA DE PROCESOS DE NEGOCIO', '4', 'II5054', 'GESTION DE LA ESTRATEGIA INDUSTRIAL', '3'],
            ['II0803', 'DISEÑO DE PRODUCTO', '3', 'II4045', 'DISEÑO DE PRODUCTO Y SERVICIOS', '3'],
            ['II0804', 'GESTIÓN DE PROYECTOS', '3', 'II3035', 'GESTIÓN DE PROYECTOS', '3'],
            ['II0805', 'DISTRIBUCIÓN Y LOCALIZACIÓN DE INSTALACIONES', '4', 'II5052', 'INGENIERIA DE INSTALACIONES Y DE ENERGÍA', '5'],
            ['II0806', 'METROLOGÍA Y NORMALIZACIÓN', '3', 'II3036', 'METROLOGIA INDUSTRIAL', '3'],
            ['SR-II', 'SEMINARIO DE REALIDAD NACIONAL II', '2', 'SR-II', 'SEMINARIO DE REALIDAD NACIONAL II', '2'],
            ['II0801', 'INGENIERÍA DE SERVICIOS', '3', 'II4043', 'INGENIERIA DE SERVICIOS', '3'],
            ['II0902', 'PROYECTO INDUSTRIAL', '3', 'II5055', 'FORMULACIÓN TFG', '2'],
            ['II0904', 'INGENIERÍA AMBIENTAL', '3', 'II3032', 'INGENIERIA DE SOSTENIBILIDAD I', '2']
        ]
        
        # Casos especiales de química que requieren combinación
        self.casos_especiales_quimica = {
            'QU0100_QU0102': {
                'cursos_requeridos': ['QU0100', 'QU0102'],
                'equivale_a': 'QU0114',
                'nombre_nuevo': 'QUIMICA GENERAL INTENSIVA'
            },
            'QU0101_QU0103': {
                'cursos_requeridos': ['QU0101', 'QU0103'],
                'equivale_a': 'QU0115',
                'nombre_nuevo': 'LABORATORIO QUIMICA GENERAL INTENSIVA'
            }
        }
        
    def analizar_expediente(self, ruta_archivo_excel: str) -> bool:
        """
        Analiza un expediente existente y agrega hoja de equiparación.
        
        Args:
            ruta_archivo_excel: Ruta al archivo Excel del expediente
            
        Returns:
            True si el análisis fue exitoso, False en caso contrario
        """
        try:
            # Abrir archivo Excel existente
            workbook = openpyxl.load_workbook(ruta_archivo_excel)
            
            # Extraer información del estudiante del nombre del archivo
            from pathlib import Path
            nombre_archivo = Path(ruta_archivo_excel).stem
            partes = nombre_archivo.split('-', 1)
            carne_estudiante = partes[0] if len(partes) > 0 else "N/A"
            nombre_estudiante = partes[1] if len(partes) > 1 else "N/A"
            
            # Verificar si ya existe la hoja de equiparación
            if 'Equiparación' in workbook.sheetnames:
                # Eliminar hoja existente para regenerar
                workbook.remove(workbook['Equiparación'])
            
            # Leer datos del historial académico
            todos_los_cursos = self._extraer_cursos_aprobados(workbook)
            
            # Realizar análisis de equiparación
            analisis = self._realizar_analisis_equiparacion(todos_los_cursos)
            
            # Crear nueva hoja de equiparación
            self._crear_hoja_equiparacion(workbook, analisis, todos_los_cursos, carne_estudiante, nombre_estudiante)
            
            # Guardar archivo
            workbook.save(ruta_archivo_excel)
            workbook.close()
            
            return True
            
        except Exception as e:
            print(f"Error analizando expediente {ruta_archivo_excel}: {str(e)}")
            return False
    
    def _extraer_cursos_aprobados(self, workbook: openpyxl.Workbook) -> Dict[str, Dict[str, Any]]:
        """
        Extrae información de TODOS los cursos del archivo Excel (no solo aprobados).
        
        Args:
            workbook: Libro de Excel abierto
            
        Returns:
            Diccionario con información de todos los cursos
        """
        todos_los_cursos = {}
        
        try:
            # Buscar hoja de expediente detallado primero
            sheet = None
            for nombre_hoja in ['Expediente Detallado', 'Historial Académico', 'Historial', 'Historial Academico']:
                if nombre_hoja in workbook.sheetnames:
                    sheet = workbook[nombre_hoja]
                    break
            
            if not sheet:
                return todos_los_cursos
            
            # Buscar encabezados en el expediente detallado
            headers = {}
            for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=10, values_only=True), 1):
                for col_idx, cell_value in enumerate(row, 1):
                    if cell_value:
                        cell_str = str(cell_value).lower()
                        if 'sigla' in cell_str:
                            headers['sigla'] = col_idx
                        elif 'curso' in cell_str and 'nombre' not in headers:
                            headers['nombre'] = col_idx
                        elif 'estado' in cell_str:
                            headers['estado'] = col_idx
                        elif 'nota' in cell_str:
                            headers['nota'] = col_idx
                        elif 'créditos' in cell_str or 'creditos' in cell_str:
                            headers['creditos'] = col_idx
                if len(headers) >= 3:  # Al menos sigla, nombre y estado
                    break
            
            if not headers.get('sigla') or not headers.get('estado'):
                return todos_los_cursos
            
            # Leer datos de cursos del expediente detallado
            start_row = 8  # Comenzar después de la información del estudiante
            for row in sheet.iter_rows(min_row=start_row, values_only=True):
                if not row or not row[headers['sigla'] - 1]:
                    continue
                
                sigla = str(row[headers['sigla'] - 1]).strip()
                if not sigla or sigla.upper() in ['SIGLA', 'SEMESTRE']:
                    continue
                
                # Extraer información del curso
                info_curso = {
                    'sigla': sigla,
                    'nombre': str(row[headers.get('nombre', 1) - 1] or '').strip(),
                    'estado': str(row[headers.get('estado', 1) - 1] or '').strip(),
                    'nota': 0.0,
                    'creditos': 0,
                    'aprobado': False
                }
                
                # Procesar nota
                if 'nota' in headers and headers['nota'] <= len(row):
                    try:
                        nota_val = row[headers['nota'] - 1]
                        if nota_val is not None and str(nota_val).strip():
                            info_curso['nota'] = float(nota_val)
                    except (ValueError, TypeError):
                        pass
                
                # Procesar créditos
                if 'creditos' in headers and headers['creditos'] <= len(row):
                    try:
                        creditos_val = row[headers['creditos'] - 1]
                        if creditos_val is not None:
                            info_curso['creditos'] = int(creditos_val)
                    except (ValueError, TypeError):
                        pass
                
                # Determinar si está aprobado basado en el estado
                estado_upper = info_curso['estado'].upper()
                if estado_upper in ['APROBADO', 'EQUIVALENTE', 'CONVALIDADO']:
                    info_curso['aprobado'] = True
                
                # Guardar curso
                todos_los_cursos[sigla] = info_curso
            
        except Exception as e:
            print(f"Error extrayendo cursos: {str(e)}")
        
        return todos_los_cursos
    
    def _realizar_analisis_equiparacion(self, cursos_aprobados: Dict[str, Dict[str, Any]]) -> AnalisisEquiparacion:
        """
        Realiza el análisis de equiparación basado en los cursos aprobados.
        
        Args:
            cursos_aprobados: Diccionario de todos los cursos
            
        Returns:
            Objeto con el resultado del análisis (simplificado, la lógica principal está en la hoja)
        """
        analisis = AnalisisEquiparacion()
        
        # Verificar química general completa (QU0100 + QU0102)
        if (('QU0100' in cursos_aprobados and cursos_aprobados['QU0100']['aprobado']) and 
            ('QU0102' in cursos_aprobados and cursos_aprobados['QU0102']['aprobado'])):
            analisis.quimica_general_completa = True
        
        # Verificar química general II completa (QU0101 + QU0103)
        if (('QU0101' in cursos_aprobados and cursos_aprobados['QU0101']['aprobado']) and 
            ('QU0103' in cursos_aprobados and cursos_aprobados['QU0103']['aprobado'])):
            analisis.quimica_intensiva_completa = True
        
        # Verificar precálculo (MA0001 en lugar de MA0125 según la tabla)
        if 'MA0001' in cursos_aprobados and cursos_aprobados['MA0001']['aprobado']:
            analisis.precalculo_aprobado = True
        
        return analisis
    
    def _crear_hoja_equiparacion(self, workbook: openpyxl.Workbook, 
                                analisis: AnalisisEquiparacion,
                                todos_los_cursos: Dict[str, Dict[str, Any]],
                                carne_estudiante: str,
                                nombre_estudiante: str) -> None:
        """
        Crea la hoja de equiparación en el archivo Excel con información completa.
        
        Args:
            workbook: Libro de Excel
            analisis: Resultado del análisis de equiparación
            todos_los_cursos: Diccionario de todos los cursos
            carne_estudiante: Carné del estudiante
            nombre_estudiante: Nombre del estudiante
        """
        # Crear nueva hoja
        sheet = workbook.create_sheet('Equiparación')
        
        # Estilos
        title_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        subheader_font = Font(name='Arial', size=10, bold=True)
        normal_font = Font(name='Arial', size=10)
        
        title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        approved_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        pending_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        no_equiv_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        equiv_fill = PatternFill(start_color='D5E8D4', end_color='D5E8D4', fill_type='solid')
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        row = 1
        
        # Información del estudiante y fecha (filas 1-3)
        sheet[f'B{row}'] = f'Estudiante: {nombre_estudiante}'
        sheet[f'B{row}'].font = Font(name='Arial', size=12, bold=True)
        row += 1
        
        sheet[f'B{row}'] = f'Carné: {carne_estudiante}'
        sheet[f'B{row}'].font = Font(name='Arial', size=12, bold=True)
        row += 1
        
        # Fecha del estudio
        from datetime import datetime
        fecha_hoy = datetime.now().strftime("%d/%m/%Y")
        sheet[f'B{row}'] = f'Estudio realizado: {fecha_hoy}'
        sheet[f'B{row}'].font = Font(name='Arial', size=11, italic=True)
        row += 2
        
        # Encabezados principales (fila 5)
        sheet.merge_cells(f'A{row}:D{row}')
        sheet[f'A{row}'] = 'PLAN DE ESTUDIOS VIGENTE'
        sheet[f'A{row}'].font = header_font
        sheet[f'A{row}'].fill = header_fill
        sheet[f'A{row}'].alignment = center_alignment
        
        sheet.merge_cells(f'E{row}:H{row}')
        sheet[f'E{row}'] = 'PLAN DE ESTUDIOS NUEVO'
        sheet[f'E{row}'].font = header_font
        sheet[f'E{row}'].fill = header_fill
        sheet[f'E{row}'].alignment = center_alignment
        row += 1
        
        # Encabezados de columnas (fila 6)
        headers = ['Sigla', 'Curso', 'Créditos', 'Estado', 'Sigla', 'Curso', 'Créditos', 'Estado']
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=row, column=col)
            cell.value = header
            cell.font = subheader_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        row += 1
        
        # Procesar cursos del expediente detallado como base
        cursos_procesados = []
        
        # Para cada curso del expediente detallado
        for sigla_vieja, curso_info in todos_los_cursos.items():
            # Buscar equivalencia en la tabla
            equivalencia_encontrada = None
            for equivalencia in self.tabla_equivalencias:
                if equivalencia[0] == sigla_vieja:  # sigla_vieja coincide
                    equivalencia_encontrada = equivalencia
                    break
            
            if equivalencia_encontrada:
                sigla_vieja, curso_viejo, creditos_viejo, sigla_nueva, curso_nuevo, creditos_nuevo = equivalencia_encontrada
            else:
                # Curso no tiene equivalencia conocida
                sigla_vieja = curso_info['sigla']
                curso_viejo = curso_info['nombre']
                creditos_viejo = curso_info['creditos']
                sigla_nueva = ''
                curso_nuevo = ''
                creditos_nuevo = 0
            
            # Estado en plan vigente (exactamente como está en el expediente)
            estado_viejo = curso_info['estado']
            fill_viejo = None
            
            if estado_viejo.upper() in ['APROBADO', 'EQUIVALENTE', 'CONVALIDADO']:
                fill_viejo = approved_fill
            elif estado_viejo.upper() == 'REPROBADO':
                fill_viejo = pending_fill
            elif estado_viejo.upper() == 'MATRICULADO':
                fill_viejo = no_equiv_fill
            
            # Estado en plan nuevo (aplicar lógica de equiparación)
            estado_nuevo = ''
            fill_nuevo = None
            
            if not sigla_nueva:  # No tiene equivalencia
                estado_nuevo = 'Sin equivalencia'
                fill_nuevo = no_equiv_fill
            elif sigla_vieja == sigla_nueva:  # Mismo curso en ambos planes
                estado_nuevo = estado_viejo  # Copiar exactamente el mismo estado
                fill_nuevo = fill_viejo
            elif estado_viejo.upper() in ['APROBADO', 'EQUIVALENTE', 'CONVALIDADO']:
                # Curso diferente pero aprobado en plan vigente
                
                # Casos especiales de química
                if sigla_vieja in ['QU0100'] and sigla_nueva in ['QU0114', 'QU-0114']:
                    if self._verificar_quimica_completa(['QU0100', 'QU0102'], todos_los_cursos):
                        estado_nuevo = 'EQUIPARADO'
                        fill_nuevo = equiv_fill
                    else:
                        estado_nuevo = 'Requiere QU0102'
                        fill_nuevo = pending_fill
                elif sigla_vieja in ['QU0102'] and sigla_nueva in ['QU0114', 'QU-0114']:
                    if self._verificar_quimica_completa(['QU0100', 'QU0102'], todos_los_cursos):
                        estado_nuevo = 'EQUIPARADO'
                        fill_nuevo = equiv_fill
                    else:
                        estado_nuevo = 'Requiere QU0100'
                        fill_nuevo = pending_fill
                elif sigla_vieja in ['QU0101'] and sigla_nueva in ['QU0115', 'QU-0115']:
                    if self._verificar_quimica_completa(['QU0101', 'QU0103'], todos_los_cursos):
                        estado_nuevo = 'EQUIPARADO'
                        fill_nuevo = equiv_fill
                    else:
                        estado_nuevo = 'Requiere QU0103'
                        fill_nuevo = pending_fill
                elif sigla_vieja in ['QU0103'] and sigla_nueva in ['QU0115', 'QU-0115']:
                    if self._verificar_quimica_completa(['QU0101', 'QU0103'], todos_los_cursos):
                        estado_nuevo = 'EQUIPARADO'
                        fill_nuevo = equiv_fill
                    else:
                        estado_nuevo = 'Requiere QU0101'
                        fill_nuevo = pending_fill
                else:
                    estado_nuevo = 'EQUIPARADO'
                    fill_nuevo = equiv_fill
            else:  # No aprobado en plan vigente
                estado_nuevo = 'Pendiente'
                fill_nuevo = pending_fill
            
            cursos_procesados.append({
                'sigla_vieja': sigla_vieja,
                'curso_viejo': curso_viejo,
                'creditos_viejo': creditos_viejo,
                'estado_viejo': estado_viejo,
                'fill_viejo': fill_viejo,
                'sigla_nueva': sigla_nueva,
                'curso_nuevo': curso_nuevo,
                'creditos_nuevo': creditos_nuevo,
                'estado_nuevo': estado_nuevo,
                'fill_nuevo': fill_nuevo
            })
        
        # Agregar cursos completamente nuevos del plan nuevo (que no están en el plan vigente)
        siglas_procesadas = set(curso['sigla_nueva'] for curso in cursos_procesados if curso['sigla_nueva'])
        
        for equivalencia in self.tabla_equivalencias:
            sigla_vieja, curso_viejo, creditos_viejo, sigla_nueva, curso_nuevo, creditos_nuevo = equivalencia
            
            # Si es un curso nuevo que no tiene equivalencia en el plan vigente
            if sigla_nueva and sigla_nueva not in siglas_procesadas and not sigla_vieja:
                cursos_procesados.append({
                    'sigla_vieja': '',
                    'curso_viejo': '',
                    'creditos_viejo': 0,
                    'estado_viejo': '',
                    'fill_viejo': None,
                    'sigla_nueva': sigla_nueva,
                    'curso_nuevo': curso_nuevo,
                    'creditos_nuevo': creditos_nuevo,
                    'estado_nuevo': 'Pendiente',
                    'fill_nuevo': pending_fill
                })
        
        # Escribir todos los cursos
        for curso in cursos_procesados:
            # Escribir datos del plan vigente
            sheet.cell(row=row, column=1).value = curso['sigla_vieja']
            sheet.cell(row=row, column=2).value = curso['curso_viejo']
            sheet.cell(row=row, column=3).value = curso['creditos_viejo']
            sheet.cell(row=row, column=4).value = curso['estado_viejo']
            
            # Aplicar formato al plan vigente
            if curso['fill_viejo']:
                for col in range(1, 5):
                    sheet.cell(row=row, column=col).fill = curso['fill_viejo']
            
            # Escribir datos del plan nuevo
            sheet.cell(row=row, column=5).value = curso['sigla_nueva']
            sheet.cell(row=row, column=6).value = curso['curso_nuevo']
            sheet.cell(row=row, column=7).value = curso['creditos_nuevo']
            sheet.cell(row=row, column=8).value = curso['estado_nuevo']
            
            # Aplicar formato al plan nuevo
            if curso['fill_nuevo']:
                for col in range(5, 9):
                    sheet.cell(row=row, column=col).fill = curso['fill_nuevo']
            
            # Aplicar fuente normal a todas las celdas
            for col in range(1, 9):
                sheet.cell(row=row, column=col).font = normal_font
            
            row += 1
        
        # Ajustar ancho de columnas
        sheet.column_dimensions['A'].width = 12  # Sigla vieja
        sheet.column_dimensions['B'].width = 45  # Curso viejo
        sheet.column_dimensions['C'].width = 10  # Créditos viejo
        sheet.column_dimensions['D'].width = 15  # Estado viejo
        sheet.column_dimensions['E'].width = 12  # Sigla nueva
        sheet.column_dimensions['F'].width = 45  # Curso nuevo
        sheet.column_dimensions['G'].width = 10  # Créditos nuevo
        sheet.column_dimensions['H'].width = 15  # Estado nuevo
        
        # Agregar leyenda explicativa
        row += 2
        sheet.cell(row=row, column=1).value = 'LEYENDA:'
        sheet.cell(row=row, column=1).font = Font(name='Arial', size=10, bold=True)
        row += 1
        
        sheet.cell(row=row, column=1).value = '• Verde claro: Curso aprobado'
        sheet.cell(row=row, column=1).font = Font(name='Arial', size=9)
        sheet.cell(row=row, column=1).fill = approved_fill
        row += 1
        
        sheet.cell(row=row, column=1).value = '• Verde oscuro: Curso equiparado'
        sheet.cell(row=row, column=1).font = Font(name='Arial', size=9)
        sheet.cell(row=row, column=1).fill = equiv_fill
        row += 1
        
        sheet.cell(row=row, column=1).value = '• Rojo: Curso pendiente o con nota insuficiente'
        sheet.cell(row=row, column=1).font = Font(name='Arial', size=9)
        sheet.cell(row=row, column=1).fill = pending_fill
        row += 1
        
        sheet.cell(row=row, column=1).value = '• Amarillo: Sin equivalencia en el nuevo plan'
        sheet.cell(row=row, column=1).font = Font(name='Arial', size=9)
        sheet.cell(row=row, column=1).fill = no_equiv_fill
    
    def _verificar_quimica_completa(self, cursos_requeridos: List[str], cursos_aprobados: Dict[str, Dict[str, Any]]) -> bool:
        """Verifica si se tienen todos los cursos requeridos para una equivalencia de química."""
        return all(curso in cursos_aprobados and cursos_aprobados[curso]['aprobado'] for curso in cursos_requeridos)
