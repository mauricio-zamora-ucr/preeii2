#!/usr/bin/env python3
"""
Script para verificar el caso específico del estudiante C14407 - problema de conteo en progreso.
"""

import openpyxl

def verificar_caso_c14407():
    """Verifica el caso específico del estudiante C14407."""
    
    archivo = './salida/C14407-ASAEL MAITA GARCIA.xlsx'
    
    try:
        workbook = openpyxl.load_workbook(archivo)
        
        print('VERIFICACIÓN CASO C14407 - PROGRESO POR SEMESTRE')
        print('=' * 55)
        
        # Verificar Progreso del Plan
        if 'Progreso del Plan' in workbook.sheetnames:
            sheet_progreso = workbook['Progreso del Plan']
            print('\nPROGRESO DEL PLAN:')
            print('-' * 20)
            
            problemas_encontrados = False
            for row_num in range(1, 50):
                semestre = sheet_progreso.cell(row=row_num, column=1).value
                if semestre and 'Semestre' in str(semestre):
                    aprobados = sheet_progreso.cell(row=row_num, column=2).value or 0
                    reprobados = sheet_progreso.cell(row=row_num, column=3).value or 0
                    matriculados = sheet_progreso.cell(row=row_num, column=4).value or 0
                    pendientes = sheet_progreso.cell(row=row_num, column=5).value or 0
                    total = sheet_progreso.cell(row=row_num, column=6).value or 0
                    
                    # Convertir valores a enteros, manejar valores None y strings
                    def to_int(value):
                        if value is None:
                            return 0
                        try:
                            return int(value)
                        except (ValueError, TypeError):
                            return 0
                    
                    aprobados_int = to_int(aprobados)
                    reprobados_int = to_int(reprobados)
                    matriculados_int = to_int(matriculados)
                    pendientes_int = to_int(pendientes)
                    total_int = to_int(total)
                    
                    suma = aprobados_int + reprobados_int + matriculados_int + pendientes_int
                    
                    print(f'   {semestre}:')
                    print(f'     Total: {total}')
                    print(f'     Aprobados: {aprobados}')
                    print(f'     Reprobados: {reprobados}')
                    print(f'     Matriculados: {matriculados}')
                    print(f'     Pendientes: {pendientes}')
                    print(f'     Suma: {suma}')
                    
                    if suma != total_int and total_int > 0:
                        print(f'     ❌ ERROR: La suma ({suma}) no coincide con el total ({total_int})')
                        problemas_encontrados = True
                    elif suma == total_int:
                        print(f'     ✅ OK: Los números cuadran correctamente')
                    
                    print()
            
            if not problemas_encontrados:
                print('🎉 No se encontraron problemas de conteo')
            else:
                print('⚠️ Se encontraron problemas de conteo que necesitan corrección')
        
        # Mostrar algunos cursos del expediente detallado para contexto
        if 'Expediente Detallado' in workbook.sheetnames:
            sheet_expediente = workbook['Expediente Detallado']
            print('\nMUESTRA DEL EXPEDIENTE DETALLADO (primeros 10 cursos):')
            print('-' * 50)
            
            count = 0
            for row_num in range(8, 200):  # Empezar después de headers
                if count >= 10:
                    break
                    
                semestre = sheet_expediente.cell(row=row_num, column=1).value
                sigla = sheet_expediente.cell(row=row_num, column=2).value
                curso = sheet_expediente.cell(row=row_num, column=3).value
                estado = sheet_expediente.cell(row=row_num, column=5).value
                
                if sigla and semestre:
                    print(f'   Sem {semestre}: {sigla} - {estado}')
                    count += 1
        
        workbook.close()
        print('\n✅ Verificación completada')
        
    except FileNotFoundError:
        print(f'✗ Error: No se encontró el archivo {archivo}')
        return False
    except Exception as e:
        print(f'✗ Error: {str(e)}')
        return False
    
    return True

if __name__ == "__main__":
    verificar_caso_c14407()
