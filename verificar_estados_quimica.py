#!/usr/bin/env python3
"""
Script para verificar específicamente los estados de química.
"""

import openpyxl

def verificar_estados_quimica():
    """Verifica los estados de los cursos de química."""
    
    try:
        # Leer archivo
        workbook = openpyxl.load_workbook('./salida/TEST-EQUIPARACION-CON-QUIMICA.xlsx')
        sheet = workbook['Equiparación']

        print('Verificación de estados de química:')
        print('=' * 60)

        # Buscar filas de química (10, 11, 18, 19)
        filas_quimica = [10, 11, 18, 19]

        for fila in filas_quimica:
            sigla_vieja = sheet.cell(row=fila, column=1).value or ''
            curso_viejo = sheet.cell(row=fila, column=2).value or ''
            estado_viejo = sheet.cell(row=fila, column=4).value or ''
            sigla_nueva = sheet.cell(row=fila, column=5).value or ''
            curso_nuevo = sheet.cell(row=fila, column=6).value or ''
            estado_nuevo = sheet.cell(row=fila, column=8).value or ''
            
            print(f'Fila {fila:2d}: {sigla_vieja:6} ({estado_viejo:8}) -> {sigla_nueva:8} ({estado_nuevo})')

        workbook.close()
        
    except Exception as e:
        print(f"Error: {str(e)}")

if __name__ == "__main__":
    verificar_estados_quimica()
