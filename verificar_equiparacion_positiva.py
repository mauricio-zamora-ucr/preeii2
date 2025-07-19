#!/usr/bin/env python3
"""
Script para verificar el análisis positivo de equiparación.
"""

import sys
sys.path.append('.')

import openpyxl

def verificar_equiparacion_positiva():
    """Verifica el análisis de equiparación para el caso positivo."""
    
    archivo_prueba = './salida/TEST-EQUIPARACION-CON-QUIMICA.xlsx'
    
    print("VERIFICACIÓN DE ANÁLISIS POSITIVO DE EQUIPARACIÓN")
    print("=" * 60)
    print(f"Archivo: {archivo_prueba}")
    print()
    
    try:
        # Abrir archivo Excel
        workbook = openpyxl.load_workbook(archivo_prueba)
        
        if 'Equiparación' in workbook.sheetnames:
            print("✓ Hoja 'Equiparación' encontrada")
            
            # Leer datos de la hoja
            sheet = workbook['Equiparación']
            
            print()
            print("Contenido de la hoja de equiparación:")
            print("-" * 60)
            
            # Leer todas las filas con contenido
            for row_num in range(1, 25):
                row_data = []
                for col_num in range(1, 7):  # Columnas A-F
                    cell = sheet.cell(row=row_num, column=col_num)
                    if cell.value:
                        row_data.append(str(cell.value))
                
                if row_data:
                    print(f"Fila {row_num:2d}: {' | '.join(row_data)}")
        else:
            print("✗ Hoja 'Equiparación' no encontrada")
        
        workbook.close()
        
    except Exception as e:
        print(f"✗ Error: {str(e)}")

if __name__ == "__main__":
    verificar_equiparacion_positiva()
