#!/usr/bin/env python3
"""
Script para verificar todos los estados en la equiparación.
"""

import openpyxl

def verificar_todos_estados():
    """Verifica todos los estados de equiparación."""
    
    archivo = './salida/C04044-BRENDA JIMENEZ MIRANDA.xlsx'
    
    try:
        workbook = openpyxl.load_workbook(archivo)
        sheet = workbook['Equiparación']

        print('VERIFICACIÓN COMPLETA DE ESTADOS EN EQUIPARACIÓN:')
        print('=' * 60)

        # Información del estudiante
        for row_num in range(1, 6):
            cell_value = sheet.cell(row=row_num, column=2).value
            if cell_value:
                print(f'  {cell_value}')

        print()
        print('ANÁLISIS COMPLETO DE ESTADOS:')
        print('-' * 40)
        print(f'{"PLAN VIGENTE":<40} {"PLAN NUEVO":<40}')
        print(f'{"Sigla":<8} {"Estado":<15} {"Sigla":<8} {"Estado":<15}')
        print('-' * 80)

        # Contar estados
        estados_vigente = {}
        estados_nuevo = {}
        equiparaciones = 0
        
        for row_num in range(7, 100):  # Revisar muchas filas
            sigla_vieja = sheet.cell(row=row_num, column=1).value
            estado_viejo = sheet.cell(row=row_num, column=4).value
            sigla_nueva = sheet.cell(row=row_num, column=5).value
            estado_nuevo = sheet.cell(row=row_num, column=8).value
            
            # Parar si no hay más datos
            if not sigla_vieja and not sigla_nueva:
                break
                
            # Mostrar algunos casos significativos
            if estado_viejo or estado_nuevo:
                sigla_vieja = sigla_vieja or ""
                estado_viejo = estado_viejo or ""
                sigla_nueva = sigla_nueva or ""
                estado_nuevo = estado_nuevo or ""
                
                print(f'{sigla_vieja:<8} {estado_viejo:<15} {sigla_nueva:<8} {estado_nuevo:<15}')
                
                # Contar estados
                if estado_viejo:
                    estados_vigente[estado_viejo] = estados_vigente.get(estado_viejo, 0) + 1
                if estado_nuevo:
                    estados_nuevo[estado_nuevo] = estados_nuevo.get(estado_nuevo, 0) + 1
                    if estado_nuevo == 'EQUIPARADO':
                        equiparaciones += 1

        print()
        print('RESUMEN DE ESTADOS:')
        print('=' * 30)
        
        print()
        print('Plan Vigente:')
        for estado, count in estados_vigente.items():
            print(f'  {estado}: {count}')
        
        print()
        print('Plan Nuevo:')
        for estado, count in estados_nuevo.items():
            print(f'  {estado}: {count}')
        
        print()
        print(f'Total equiparaciones detectadas: {equiparaciones}')
        
        workbook.close()
        
        print()
        print('✅ Verificación completada exitosamente')
        
    except Exception as e:
        print(f'✗ Error: {str(e)}')

if __name__ == "__main__":
    verificar_todos_estados()
