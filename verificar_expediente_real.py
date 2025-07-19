#!/usr/bin/env python3
"""
Script para verificar expediente real mejorado.
"""

import openpyxl

def verificar_expediente_real():
    """Verifica el expediente real con las mejoras."""
    
    archivo = './salida/C04044-BRENDA JIMENEZ MIRANDA.xlsx'
    
    try:
        workbook = openpyxl.load_workbook(archivo)
        sheet = workbook['Equiparación']

        print('VERIFICACIÓN DE EXPEDIENTE REAL:')
        print('=' * 50)

        # Información del estudiante
        for row_num in range(1, 6):
            cell_value = sheet.cell(row=row_num, column=2).value
            if cell_value:
                print(f'  {cell_value}')

        print()
        print('MUESTRA DE CURSOS CON ESTADOS:')
        print('-' * 35)

        # Buscar algunos cursos con estados interesantes
        count = 0
        for row_num in range(7, 60):  # Revisar más filas
            sigla_vieja = sheet.cell(row=row_num, column=1).value
            estado_viejo = sheet.cell(row=row_num, column=4).value
            sigla_nueva = sheet.cell(row=row_num, column=5).value
            estado_nuevo = sheet.cell(row=row_num, column=8).value
            
            # Mostrar solo filas con estados relevantes
            if estado_viejo or estado_nuevo:
                print(f'{sigla_vieja or "":<8} ({estado_viejo or "":<12}) -> {sigla_nueva or "":<8} ({estado_nuevo or ""})')
                count += 1
                if count >= 10:  # Limitar salida
                    break

        workbook.close()
        
        print()
        print('✅ Verificación completada exitosamente')
        
    except Exception as e:
        print(f'✗ Error: {str(e)}')

if __name__ == "__main__":
    verificar_expediente_real()
