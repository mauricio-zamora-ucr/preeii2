#!/usr/bin/env python3
"""
Script para verificar que los archivos Excel generados contengan las 3 hojas correctas
"""
import os
import glob
try:
    import openpyxl
except ImportError:
    print("Instalando openpyxl para verificar archivos Excel...")
    os.system("pip install openpyxl")
    import openpyxl

def verificar_archivo_excel(ruta_archivo):
    """Verifica que un archivo Excel contenga las 3 hojas esperadas"""
    try:
        workbook = openpyxl.load_workbook(ruta_archivo)
        hojas = workbook.sheetnames
        print(f"\nArchivo: {os.path.basename(ruta_archivo)}")
        print(f"Hojas encontradas: {hojas}")
        
        hojas_esperadas = ['Malla Curricular', 'Expediente Detallado', 'Historial Completo']
        for hoja in hojas_esperadas:
            if hoja in hojas:
                print(f"  ✓ {hoja}")
            else:
                print(f"  ✗ {hoja} - NO ENCONTRADA")
        
        # Verificar contenido de una hoja
        if 'Malla Curricular' in hojas:
            sheet = workbook['Malla Curricular']
            print(f"  - Malla Curricular tiene {sheet.max_row} filas y {sheet.max_column} columnas")
        
        workbook.close()
        return True
        
    except Exception as e:
        print(f"Error al abrir {ruta_archivo}: {str(e)}")
        return False

def main():
    """Función principal"""
    print("Verificando archivos Excel generados...")
    
    # Buscar archivos Excel en la carpeta salida
    patron = "/workspaces/preeii2/salida/*.xlsx"
    archivos = glob.glob(patron)
    
    if not archivos:
        print("No se encontraron archivos Excel en la carpeta salida/")
        return
    
    print(f"Se encontraron {len(archivos)} archivos Excel")
    
    # Verificar solo los primeros 3 archivos para no saturar la salida
    for archivo in archivos[:3]:
        verificar_archivo_excel(archivo)
    
    if len(archivos) > 3:
        print(f"\n... y {len(archivos) - 3} archivos más")

if __name__ == "__main__":
    main()
