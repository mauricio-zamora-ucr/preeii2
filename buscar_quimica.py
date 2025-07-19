#!/usr/bin/env python3
"""
Script para buscar expedientes con cursos de química aprobados.
"""

import sys
sys.path.append('.')

import openpyxl
from pathlib import Path

def buscar_cursos_quimica():
    """Busca expedientes que tengan cursos de química aprobados."""
    
    print("BÚSQUEDA DE EXPEDIENTES CON CURSOS DE QUÍMICA")
    print("=" * 60)
    print()
    
    directorio_salida = Path('./salida')
    archivos_excel = list(directorio_salida.glob('*.xlsx'))
    archivos_excel = [f for f in archivos_excel if not f.name.startswith('TEST-')]
    
    cursos_quimica = ['QU0100', 'QU0101', 'QU0102', 'QU0103']
    curso_precalculo = 'MA0125'
    
    resultados = []
    
    for archivo in archivos_excel[:10]:  # Revisar solo los primeros 10 para prueba
        try:
            workbook = openpyxl.load_workbook(str(archivo))
            
            # Buscar hoja de historial
            sheet = None
            for nombre_hoja in ['Historial Académico', 'Historial', 'Historial Academico']:
                if nombre_hoja in workbook.sheetnames:
                    sheet = workbook[nombre_hoja]
                    break
            
            if not sheet:
                workbook.close()
                continue
            
            # Buscar cursos
            cursos_encontrados = []
            
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                
                codigo = str(row[0]).strip()
                if codigo in cursos_quimica or codigo == curso_precalculo:
                    # Verificar si está aprobado
                    nota = 0.0
                    resultado = ''
                    
                    # Intentar leer nota (normalmente en columna 3 o 4)
                    for col_idx in [2, 3, 4]:
                        if col_idx < len(row) and row[col_idx] is not None:
                            try:
                                nota = float(row[col_idx])
                                break
                            except (ValueError, TypeError):
                                pass
                    
                    # Intentar leer resultado
                    for col_idx in range(len(row)):
                        if row[col_idx] and str(row[col_idx]).upper() in ['APR', 'APROBADO', 'A']:
                            resultado = str(row[col_idx])
                            break
                    
                    if nota >= 7.0 or resultado.upper() in ['APR', 'APROBADO', 'A']:
                        cursos_encontrados.append(codigo)
            
            if cursos_encontrados:
                nombre_archivo = archivo.stem
                partes = nombre_archivo.split('-', 1)
                carne = partes[0] if len(partes) > 0 else "N/A"
                nombre = partes[1] if len(partes) > 1 else "N/A"
                
                resultados.append({
                    'archivo': archivo.name,
                    'carne': carne,
                    'nombre': nombre,
                    'cursos': cursos_encontrados
                })
            
            workbook.close()
            
        except Exception as e:
            print(f"Error procesando {archivo.name}: {str(e)}")
    
    # Mostrar resultados
    if resultados:
        print("Expedientes con cursos de química/precálculo aprobados:")
        print("-" * 60)
        for resultado in resultados:
            print(f"Archivo: {resultado['archivo']}")
            print(f"Estudiante: {resultado['carne']} - {resultado['nombre']}")
            print(f"Cursos aprobados: {', '.join(resultado['cursos'])}")
            print()
    else:
        print("No se encontraron expedientes con cursos de química/precálculo aprobados")

if __name__ == "__main__":
    buscar_cursos_quimica()
