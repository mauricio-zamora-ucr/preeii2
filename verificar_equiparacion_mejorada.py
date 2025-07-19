#!/usr/bin/env python3
"""
Script para verificar la hoja de equiparación mejorada.
"""

import openpyxl

def verificar_equiparacion_mejorada():
    """Verifica la nueva estructura de la hoja de equiparación."""
    
    archivo_prueba = './salida/TEST-EQUIPARACION-CON-QUIMICA.xlsx'
    
    print("VERIFICACIÓN DE HOJA DE EQUIPARACIÓN MEJORADA")
    print("=" * 60)
    print(f"Archivo: {archivo_prueba}")
    print()
    
    try:
        # Abrir archivo Excel
        workbook = openpyxl.load_workbook(archivo_prueba)
        
        if 'Equiparación' in workbook.sheetnames:
            print("✓ Hoja 'Equiparación' encontrada")
            
            # Leer datos de la hoja
            sheet = workbook['Equiparación']
            
            print()
            print("INFORMACIÓN DEL ESTUDIANTE:")
            print("-" * 30)
            
            # Leer información del estudiante
            for row_num in range(1, 6):
                cell_value = sheet.cell(row=row_num, column=2).value
                if cell_value:
                    print(f"  {cell_value}")
            
            print()
            print("ESTRUCTURA DE LA TABLA:")
            print("-" * 25)
            
            # Leer encabezados
            for row_num in range(5, 8):
                row_data = []
                for col_num in range(1, 9):
                    cell = sheet.cell(row=row_num, column=col_num)
                    if cell.value:
                        row_data.append(str(cell.value))
                    else:
                        row_data.append('')
                
                if any(data.strip() for data in row_data):
                    print(f"Fila {row_num}: {' | '.join(row_data)}")
            
            print()
            print("MUESTRA DE DATOS (PRIMERAS 5 FILAS):")
            print("-" * 40)
            
            # Leer algunas filas de datos
            for row_num in range(8, 13):
                row_data = []
                for col_num in range(1, 9):
                    cell = sheet.cell(row=row_num, column=col_num)
                    if cell.value:
                        row_data.append(str(cell.value)[:20])  # Truncar nombres largos
                    else:
                        row_data.append('')
                
                if any(data.strip() for data in row_data):
                    print(f"Fila {row_num}: {' | '.join(row_data)}")
            
        else:
            print("✗ Hoja 'Equiparación' no encontrada")
        
        workbook.close()
        
    except Exception as e:
        print(f"✗ Error: {str(e)}")

if __name__ == "__main__":
    verificar_equiparacion_mejorada()
