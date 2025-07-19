#!/usr/bin/env python3
"""
Script para verificar que la hoja de equiparación se creó correctamente.
"""

import sys
sys.path.append('.')

import openpyxl

def verificar_hoja_equiparacion():
    """Verifica que la hoja de equiparación se creó correctamente."""
    
    archivo_prueba = './salida/C04044-BRENDA JIMENEZ MIRANDA.xlsx'
    
    print("VERIFICACIÓN DE HOJA DE EQUIPARACIÓN")
    print("=" * 50)
    print(f"Archivo: {archivo_prueba}")
    print()
    
    try:
        # Abrir archivo Excel
        workbook = openpyxl.load_workbook(archivo_prueba)
        
        print("Hojas disponibles en el archivo:")
        for i, sheet_name in enumerate(workbook.sheetnames, 1):
            print(f"  {i}. {sheet_name}")
        
        # Verificar si existe la hoja de equiparación
        if 'Equiparación' in workbook.sheetnames:
            print()
            print("✓ Hoja 'Equiparación' encontrada")
            
            # Leer algunos datos de la hoja
            sheet = workbook['Equiparación']
            
            print()
            print("Contenido de la hoja de equiparación:")
            print("-" * 40)
            
            # Leer las primeras 15 filas
            for row_num in range(1, 16):
                row_data = []
                for col_num in range(1, 7):  # Columnas A-F
                    cell = sheet.cell(row=row_num, column=col_num)
                    if cell.value:
                        row_data.append(str(cell.value))
                
                if row_data:
                    print(f"Fila {row_num:2d}: {' | '.join(row_data)}")
        else:
            print("✗ Hoja 'Equiparación' no encontrada")
        
        workbook.close()
        
    except Exception as e:
        print(f"✗ Error: {str(e)}")

if __name__ == "__main__":
    verificar_hoja_equiparacion()
