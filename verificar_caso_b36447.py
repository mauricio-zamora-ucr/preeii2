#!/usr/bin/env python3
"""
Script para verificar el caso específico del estudiante B36447 - MA1004 ÁLGEBRA LINEAL.
"""

import openpyxl
import sys

def verificar_caso_b36447():
    """Verifica el caso específico del estudiante B36447."""
    
    archivo = './salida/B36447-ARIEL SANCHEZ MONTEALEGRE.xlsx'
    
    try:
        workbook = openpyxl.load_workbook(archivo)
        
        print('VERIFICACIÓN CASO B36447 - MA1004 ÁLGEBRA LINEAL')
        print('=' * 60)
        
        # Verificar en el Historial Completo
        if 'Historial Completo' in workbook.sheetnames:
            sheet_historial = workbook['Historial Completo']
            print('\n1. HISTORIAL COMPLETO:')
            print('-' * 25)
            
            found_ma1004 = False
            for row_num in range(1, 200):  # Revisar muchas filas
                sigla = sheet_historial.cell(row=row_num, column=1).value
                if sigla == 'MA1004':
                    found_ma1004 = True
                    curso = sheet_historial.cell(row=row_num, column=2).value
                    estado = sheet_historial.cell(row=row_num, column=7).value
                    nota = sheet_historial.cell(row=row_num, column=8).value
                    anno = sheet_historial.cell(row=row_num, column=6).value
                    periodo = sheet_historial.cell(row=row_num, column=5).value
                    print(f'   {sigla} - {curso}: {estado} (Nota: {nota}) [{anno}-{periodo}]')
            
            if not found_ma1004:
                print('   MA1004 no encontrado en historial completo')
        
        # Verificar en el Expediente Detallado
        if 'Expediente Detallado' in workbook.sheetnames:
            sheet_expediente = workbook['Expediente Detallado']
            print('\n2. EXPEDIENTE DETALLADO:')
            print('-' * 25)
            
            found_ma1004 = False
            for row_num in range(1, 200):
                sigla = sheet_expediente.cell(row=row_num, column=2).value
                if sigla == 'MA1004':
                    found_ma1004 = True
                    curso = sheet_expediente.cell(row=row_num, column=3).value
                    estado = sheet_expediente.cell(row=row_num, column=5).value
                    nota = sheet_expediente.cell(row=row_num, column=6).value
                    print(f'   {sigla} - {curso}: {estado} (Nota: {nota}) [ESTADO PRINCIPAL]')
            
            if not found_ma1004:
                print('   MA1004 no encontrado en expediente detallado')
        
        # Verificar en Progreso del Plan
        if 'Progreso del Plan' in workbook.sheetnames:
            sheet_progreso = workbook['Progreso del Plan']
            print('\n3. PROGRESO DEL PLAN:')
            print('-' * 20)
            
            for row_num in range(1, 50):
                semestre = sheet_progreso.cell(row=row_num, column=1).value
                if semestre and 'Semestre' in str(semestre):
                    aprobados = sheet_progreso.cell(row=row_num, column=2).value
                    reprobados = sheet_progreso.cell(row=row_num, column=3).value
                    matriculados = sheet_progreso.cell(row=row_num, column=4).value
                    pendientes = sheet_progreso.cell(row=row_num, column=5).value
                    total = sheet_progreso.cell(row=row_num, column=6).value
                    
                    # Convertir valores a enteros, manejar valores None y strings
                    def to_int(value):
                        if value is None:
                            return 0
                        try:
                            return int(value)
                        except (ValueError, TypeError):
                            return 0
                    
                    aprobados_int = to_int(aprobados)
                    reprobados_int = to_int(reprobados)
                    matriculados_int = to_int(matriculados)
                    pendientes_int = to_int(pendientes)
                    total_int = to_int(total)
                    
                    suma = aprobados_int + reprobados_int + matriculados_int + pendientes_int
                    
                    print(f'   {semestre}: Total={total}, Aprobados={aprobados}, Reprobados={reprobados}, Matriculados={matriculados}, Pendientes={pendientes}')
                    
                    if suma != total_int and total_int > 0:
                        print(f'     ⚠️  ERROR: La suma ({suma}) no coincide con el total ({total_int})')
        
        workbook.close()
        print('\n✅ Verificación completada')
        
    except FileNotFoundError:
        print(f'✗ Error: No se encontró el archivo {archivo}')
        return False
    except Exception as e:
        print(f'✗ Error: {str(e)}')
        return False
    
    return True

if __name__ == "__main__":
    verificar_caso_b36447()
