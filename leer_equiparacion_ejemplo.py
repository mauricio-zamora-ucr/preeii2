#!/usr/bin/env python3
"""
Script para leer el archivo Equiparacion.xlsx de ejemplo.
"""

import openpyxl
import sys

def leer_equiparacion_ejemplo():
    """Lee el archivo de ejemplo para ver la estructura correcta."""
    
    try:
        # Abrir el archivo de ejemplo
        workbook = openpyxl.load_workbook('/workspaces/preeii2/Equiparacion.xlsx')
        print('Hojas disponibles:', workbook.sheetnames)
        print()

        # Leer la primera hoja
        sheet = workbook.active
        print('Contenido del archivo Equiparacion.xlsx:')
        print('=' * 100)

        # Leer todas las filas con contenido
        for row_num in range(1, 60):  # Leer hasta fila 60
            row_data = []
            for col_num in range(1, 12):  # Columnas A-K
                cell = sheet.cell(row=row_num, column=col_num)
                if cell.value is not None:
                    row_data.append(str(cell.value))
                else:
                    row_data.append('')
            
            # Solo mostrar filas que tienen algún contenido
            if any(data.strip() for data in row_data):
                # Formatear para mostrar columnas claramente
                cols_str = []
                for i, data in enumerate(row_data[:8]):
                    cols_str.append(f"Col{chr(65+i)}:'{data}'" if data.strip() else f"Col{chr(65+i)}:''")
                print(f'Fila {row_num:2d}: {" | ".join(cols_str)}')

        workbook.close()
        
    except Exception as e:
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    leer_equiparacion_ejemplo()
